#!/usr/bin/env python3

from bescheinigungs_check import BescheinigungsCheck
import openpyxl
import datetime
import argparse
import logging
from openpyxl.utils.exceptions import InvalidFileException

# Check of EfZ-Nachweis via DPSG NaMi 2.2 web interface
# data from a Microsoft Excel spreadsheet is processed and updated


def get_cols_from_cli_args(cli_args):
    # dependent on the specific file structure the data is spread across different columns;
    # we need to define the column number mapping here;
    # we need first name, surname, EfZ number and birthday to start the query
    # status and update timestamp are used to store the response in the same document
    # TODO: somehow apply auto-detection from the given input file (nice to have)
    col_firstname = cli_args.columns[0]
    col_surname = cli_args.columns[1]
    col_birthdate = cli_args.columns[2]
    col_fznumber = cli_args.columns[3]
    col_status = None
    if len(cli_args.columns) > 4:
        col_status = cli_args.columns[4]
    col_update_timestamp = None
    if len(cli_args.columns) > 5:
        col_update_timestamp = cli_args.columns[5]
    return col_firstname, col_surname, col_fznumber, col_birthdate, col_status, col_update_timestamp


def get_query_input_from_sheet_row(cli_args, sheet, row_idx):
    # get column numbers for required input data; ignore columns for optional output data (status and timestamp)
    col_firstname, col_surname, col_fznumber, col_birthdate, _, _ = get_cols_from_cli_args(cli_args)

    birthdate = sheet.cell(row=row_idx, column=col_birthdate).value
    if birthdate:
        if type(birthdate) == datetime.datetime:
            birthdate = birthdate.strftime("%d.%m.%Y")
    else:
        birthdate = ""

    fznumber = sheet.cell(row=row_idx, column=col_fznumber).value
    if not fznumber:
        fznumber = ""

    firstname = sheet.cell(row=row_idx, column=col_firstname).value
    if not firstname:
        firstname = ""

    surname = sheet.cell(row=row_idx, column=col_surname).value
    if not surname:
        surname = ""

    request_data = {
        'fzNummer': fznumber,
        'vorname': firstname,
        'nachname': surname,
        'geburtsdatum': birthdate
    }

    return request_data


def process_excel_file(cli_args):
    filename = cli_args.filename

    # TODO: we could also make this CSV input format compatible
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    log.debug(f"Processing file '{filename}'")

    start_row = cli_args.start_row  # skip the header row as it usually contains the headings
    log.debug(f"Start row: {start_row}")
    if start_row < 1:
        exit(f"Start row number may not be < 1, given: {start_row}.")
    log.info(f"Number of columns in workbook sheet: {sheet.max_column}")
    log.debug(f"Number of rows in workbook sheet to be processed: {sheet.max_row}")
    if any(col_idx > sheet.max_column for col_idx in cli_args.columns):
        log.error("Given column number exceeds number of available columns.")
        exit()

    col_firstname, col_surname, col_fznumber, col_birthdate, col_status, col_update_timestamp = \
        get_cols_from_cli_args(cli_args)
    log.debug(f"First name is taken from column #{col_firstname}, surname from #{col_surname}, "
              f"EfZ ID from #{col_fznumber} and date of birth from #{col_birthdate}.")
    could_or_will = "will"
    if cli_args.dry_run:
        could_or_will = "could"
    if col_status:
        log.info(f"Status {could_or_will} be stored in column #{col_status}.")
    else:
        log.info("Won't store status (no valid column number given).")
    if col_update_timestamp:
        log.info(f"Timestamp {could_or_will} be stored in column #{col_update_timestamp}.")
    else:
        log.info("Won't store timestamp (no valid column number given).")

    store_workbook = not cli_args.dry_run and ((col_status is not None) or (col_update_timestamp is not None))
    log.debug(f"Need to store workbook after processing: {store_workbook}")

    # print header for result visualization
    if not cli_args.dont_print:
        output_format = f"{' '*22}{'ID':<4}  {'Vorname':>20} {'Nachname':>20}   Status"
        print(output_format)
        print('-'*100)

    for row_idx in range(start_row, sheet.max_row + 1):
        req_data = get_query_input_from_sheet_row(cli_args, sheet, row_idx=row_idx)
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        efz_status = 'Prüfung fehlerhaft'
        try:
            check = BescheinigungsCheck(req_data['vorname'], req_data['nachname'], req_data['geburtsdatum'],
                                        req_data['fzNummer'])
            status = check.get_check_status()
            if status:
                efz_status = 'Gültig'
            else:
                efz_status = 'Ungültig'
        except ValueError:
            pass

        # visualize result (along with some input data) on the command line
        if not cli_args.dont_print:
            output_format = f"[{timestamp}] #{row_idx-start_row+1:<4} {req_data.get('vorname'):>20} " \
                            f"{req_data.get('nachname'):>20}   {efz_status}"
            print(output_format)

        # write result back to the workbook
        if store_workbook and col_status:
            sheet.cell(row=row_idx, column=col_status).value = efz_status
        if store_workbook and col_update_timestamp:
            sheet.cell(row=row_idx, column=col_update_timestamp).value = timestamp

    # feed back to Excel
    if store_workbook:
        wb.save(filename)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='%(prog)s checks EfZ-Nachweise in DPSG NaMi 2.2 using Excel and '
                                                 'HTTP requests')

    parser.add_argument(dest='filename',
                        help='path to Excel input/output file',
                        default="logic.json")

    parser.add_argument('-c', '--columns',
                        type=int,
                        nargs='+',
                        help='column numbers for the Excel workbook in the order '
                             'Vorname (prename), '
                             'Nachname (surname), '
                             'Geburtsdatum (date of birth), '
                             'EfZ ID (EfZ-Nummer), '
                             'Status and Timestamp '
                             'where the first column has number 1 (not 0)',
                        default=[1, 2, 3, 4, 5, 6])  # default falls back to listed order from the help text above

    parser.add_argument('-n', '--dry-run',
                        action='store_true',
                        help='Send the HTTP requests to the server but '
                             'don''t write back to the Excel file.',
                        default=False)

    parser.add_argument('-dp', '--dont-print',
                        action='store_true',
                        help='Don''t print the data on the standard output.',
                        default=False)

    parser.add_argument('-sr', '--start-row',
                        type=int,
                        help='Start row (to skip parsing of header lines); the first row has number 1 (not 0).',
                        default=2)

    parser.add_argument('-v', '--verbose',
                        action='count',
                        help='log level (-v: INFO, -vv: DEBUG)',
                        default=0)

    args = parser.parse_args()

    # Create and configure logger object
    log = logging.getLogger(__name__)

    if args.verbose > 0:
        log.setLevel(logging.INFO)
        if args.verbose > 1:
            log.setLevel(logging.DEBUG)
    else:
        log.setLevel(logging.WARNING)

    # pre-check validity of given column numbers
    num_cols = len(args.columns)
    if num_cols < 4 or num_cols > 6:
        log.error(f"Number of columns must be 4..6, but {num_cols} given: {args.columns}.")
        exit()
    if not all(col_idx > 0 for col_idx in args.columns):
        log.error(f"Only column numbers greater or equal 1 allowed, but given: {args.columns}.")
        exit()
    if len(set(args.columns)) < num_cols:
        # check for duplicates (convert to set and check length against length of list)
        log.error("Cannot use the same column number for multiple input fields.")
        exit()
    if num_cols == 4:
        log.info("Only 4 column numbers given. Won't write back status and timestamp to Excel file (read-only).")

    try:
        process_excel_file(args)
    except FileNotFoundError:
        log.error(f"Could not find file '{args.filename}'.")
        exit()
    except InvalidFileException:
        log.error(f"File format of file '{args.filename}' is currently not supported "
                  "(supported formats are .xlsx, .xlsm, .xltx and .xltm).")
        exit()
